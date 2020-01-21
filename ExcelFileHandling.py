'''
Created on 22-06-2019
@author Neeru Rani
'''

import openpyxl
import pandas as pd

#................... Method to Open an excel file....................
def openExcel(filePath):
    wb_obj = openpyxl.load_workbook(filePath)
    sheet_obj = wb_obj.active
    return wb_obj, sheet_obj

#................... Method to print all data of excel file....................
def readExcel(fpath):
    w_obj, st_obj =openExcel(fpath)
    m_row = st_obj.max_row
    m_col = st_obj.max_column
    for i in range(1, m_row + 1):
        for j in range(1, m_col + 1):
            cell_obj = st_obj.cell(row=i, column=j)
            print(cell_obj.value, end=" ")
        print()

#path = "C:\\neeru\\Python_Practice\\ItemDetails.xlsx"
#ReadExcel(path)

#................... Method to insert data in excel file....................
def insertInExcel(fpath):
    w_obj, st_obj = openExcel(fpath)
    m_row = st_obj.max_row
    m_col = st_obj.max_column
    for j in range(1, m_col + 1):
        cell_obj = st_obj.cell(row=m_row+1, column=j)
        c1=st_obj.cell(row=1, column=j)
        cell_obj.value=input("Enter " + c1.value + " value:- ")
    w_obj.save(fpath)

'''path = "C:\\neeru\\Python_Practice\\ItemDetails.xlsx"
nor = int(input("Enter no. of row that u want to add: "))
for i in range(0,nor):
    insertInExcel(path)
readExcel(path)'''

#................... Method to insert data in excel file (Without duplicate value) ....................
def insertInExcel2(fpath):
    # 1st column should be unique
    w_obj, st_obj = openExcel(fpath)
    m_row = st_obj.max_row
    m_col = st_obj.max_column
    flag = True

    c1 = st_obj.cell(row=1, column=1)
    enteredVal = input("Enter " + c1.value + " value:- ")
    for i in range(1,m_row+1):
        c2 = st_obj.cell(row=i, column=1)
        if c2.value == enteredVal:
            flag = False
    if flag == True:
        cell_obj = st_obj.cell(row=m_row + 1, column=1)
        cell_obj.value = enteredVal
        for j in range(2, m_col + 1):
            cell_obj = st_obj.cell(row=m_row + 1, column=j)
            c3 = st_obj.cell(row=1, column=j)
            cell_obj.value = input("Enter " + c3.value + " value:- ")
    else:
          print("Entered value is already exist...")

    w_obj.save(fpath)

'''path = "C:\\neeru\\Python_Practice\\ItemDetails.xlsx"
insertInExcel2(path)
#readExcel(path)'''

#................... Method to update data in excel file....................
def updateExcel(uniqueColVal,colNameToUpdate,newVal,fpath):
    #uniqueColVal  -->  Unique value to match (of which row's column's value want to change)
    #colNameToUpdate --> column name (of which particular column's value want to change)
    w_obj, st_obj = openExcel(fpath)
    flag = False

    for i in range(1, st_obj.max_row+1):
        for j in range(1, st_obj.max_column + 1):
            cell_obj = st_obj.cell(row=i, column=j)
            if cell_obj.value == uniqueColVal:
                flag = True
                for col in range(1, st_obj.max_column + 1):
                    c1 = st_obj.cell(row=1, column=col)
                    if (colNameToUpdate == c1.value):
                        changeing_pos = st_obj.cell(row=i, column=col)
                        changeing_pos.value=newVal
                break
    if flag == False:
        print("Entered value is not available to update")
    w_obj.save(fpath)

'''path = "C:\\neeru\\Python_Practice\\ItemDetails.xlsx"
updateExcel("ss","DateOfMfd",500,path)
readExcel(path)'''

#................... Method to delete data from excel file....................
def deleteFromExcel(uniqueColName,rowVal,fpath):
    w_obj, st_obj = openExcel(fpath)
    m_row = st_obj.max_row
    m_col = st_obj.max_column
    flag = False

    for col in range(1,m_col + 1):
        c1 = st_obj.cell(row=1, column=col)
        if c1.value == uniqueColName:

            for i in range(2, m_row + 1):
                for j in range(1, m_col + 1):
                    cell_obj = st_obj.cell(row=i, column=j)
                    if cell_obj.value == rowVal:
                        for a in range(1, m_col + 1):
                            cell_obj = st_obj.cell(row=i, column=a)
                            cell_obj.value = " "
                        for a in range(i,m_row + 1):
                            for b in range(1, m_col + 1):
                                upper_cell = st_obj.cell(row=a, column=b)
                                lower_cell = st_obj.cell(row=a+1, column=b)
                                upper_cell.value = lower_cell.value
                        flag = True
                        break
    if flag == False:
        print("Entered value is not available to Delete")
    else:
        print("Entered row is Deleted")
    w_obj.save(fpath)


'''path = "C:\\neeru\\Python_Practice\\ItemDetails.xlsx"
deleteFromExcel("Item","ss",path)
readExcel(path)'''





